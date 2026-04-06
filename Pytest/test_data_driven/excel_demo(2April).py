import openpyxl

# creating workbook
wb = openpyxl.Workbook()

# creating sheet
sheetName = "Sheet1"

if sheetName in wb.sheetnames:
    ws = wb[sheetName] # if sheet already present, fetch in current working sheet
else:
    ws = wb.create_sheet(sheetName) # if sheet not present create sheet

# entering data
ws['A1'] = 'User'
ws['B1'] = 'Password'

# saving workbook
#wb.save('sample.xlsx') #will be saved in the local directory

#inserting the data
ws.append(['user1','1232'])
ws.append(['user2','1234'])
ws.append(['user3','1239'])
ws.append(['user4','1230'])
wb.save('sample.xlsx')

# fetching the data
""" print(ws['A2'].value)
print(ws['B2'].value)
print(ws['A3'].value)
print(ws['B3'].value) """

# To iterate in the excel sheet
for row in ws.iter_rows(values_only=True):
    print(row[0],row[1])



